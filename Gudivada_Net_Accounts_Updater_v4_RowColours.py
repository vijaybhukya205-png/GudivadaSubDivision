import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BO_TARGET = 300
SO_TARGET = 600
HO_TARGET = 2000

def choose_file(title):
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    p = filedialog.askopenfilename(
        title=title,
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("CSV files", "*.csv"),
            ("All files", "*.*")
        ]
    )
    root.destroy()
    return Path(p) if p else None

def read_cept_file(path):
    raw = pd.read_excel(path, header=None)

    # Find the real header row by looking for Name + SO Name.
    header = None
    for i in range(min(15, len(raw))):
        row = raw.iloc[i].astype(str).str.strip().str.lower()
        if "name" in row.values and "so name" in row.values:
            header = i
            break

    if header is None:
        raise ValueError(
            f"Could not identify CEPT header row in {path.name}. "
            "The report must contain 'Name' and 'SO Name' headers."
        )

    df = raw.iloc[header + 1:].copy()
    df.columns = [str(x).strip() for x in raw.iloc[header].tolist()]

    # Remove unnamed/blank columns.
    df = df[
        [c for c in df.columns
         if str(c).strip() and not str(c).lower().startswith("unnamed")]
    ]

    # Locate Name column robustly.
    name_col = next(
        (c for c in df.columns if str(c).strip().lower() == "name"),
        None
    )

    if name_col is None:
        raise ValueError(
            f"'Name' column not found in {path.name}. "
            f"Detected columns: {list(df.columns)}"
        )

    # Remove CEPT total row and blank rows.
    df[name_col] = df[name_col].astype(str).str.strip()
    df = df[
        (df[name_col] != "") &
        (df[name_col].str.lower() != "nan") &
        (df[name_col].str.lower() != "total:")
    ].copy()

    return df, name_col

def calculate_report(path, exclude_last):
    df, name_col = read_cept_file(path)

    # CEPT report structure:
    # Name | SOL ID / BOCODE | SO Name | product columns...
    metadata = {
        str(c).strip().lower()
        for c in df.columns
        if str(c).strip().lower() in {
            "name", "sol id / bocode", "so name"
        }
    }

    product_cols = [
        c for c in df.columns
        if str(c).strip().lower() not in metadata
    ]

    if len(product_cols) <= exclude_last:
        raise ValueError(
            f"{path.name}: only {len(product_cols)} product columns found; "
            f"cannot exclude the last {exclude_last}."
        )

    # User's rule:
    # Opened = all product columns except last 4
    # Closed = all product columns except last 6
    product_cols = product_cols[:-exclude_last]

    values = df[product_cols].apply(
        pd.to_numeric, errors="coerce"
    ).fillna(0).sum(axis=1)

    result = pd.DataFrame({
        "Office Name": df[name_col].astype(str).str.strip(),
        "Value": values
    })

    return result.groupby(
        "Office Name", as_index=False
    )["Value"].sum()

def office_type(name):
    n = str(name).upper().strip()

    # CEPT naming examples: X B.O, X S.O, X H.O
    if " H.O" in n or n.endswith("H.O"):
        return "HO"
    if " S.O" in n or n.endswith("S.O"):
        return "SO"
    return "BO"

def build_dashboard(opened_path, closed_path):
    opened = calculate_report(opened_path, 4)
    closed = calculate_report(closed_path, 6)

    opened = opened.rename(columns={"Value": "Accounts Opened"})
    closed = closed.rename(columns={"Value": "Accounts Closed"})

    data = opened.merge(
        closed,
        on="Office Name",
        how="outer"
    ).fillna(0)

    data["Net Accounts"] = (
        data["Accounts Opened"] - data["Accounts Closed"]
    )

    data["Office Type"] = data["Office Name"].map(office_type)

    data["Annual Target"] = data["Office Type"].map({
        "BO": BO_TARGET,
        "SO": SO_TARGET,
        "HO": HO_TARGET
    })

    # FY proportionate target including the current month.
    month = date.today().month
    fy_month = month - 3 if month >= 4 else month + 9

    data["Proportionate Target"] = (
        data["Annual Target"] * fy_month / 12
    )

    data["Achievement %"] = (
        data["Net Accounts"] /
        data["Proportionate Target"] * 100
    ).replace([float("inf"), -float("inf")], 0).fillna(0)

    data["Yet to Achieve"] = (
        data["Proportionate Target"] -
        data["Net Accounts"]
    ).clip(lower=0)

    return data

def style_workbook(path):
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    total_font = Font(bold=True)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    green_font = Font(color="006100", bold=True)
    red_font = Font(color="9C0006", bold=True)
    thin = Side(style="thin", color="D9E1F2")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)

        widths = {"A":8, "B":32, "C":14, "D":18, "E":18,
                  "F":16, "G":16, "H":22, "I":18, "J":22}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}
        ach_col = headers.get("Achievement %")
        yet_col = headers.get("Yet to Achieve")
        net_col = headers.get("Net Accounts")

        if ach_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, ach_col)
                cell.number_format = '0.0"%"'

                # Colour the complete office-performance row based on Achievement %.
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 100:
                        row_fill, row_font = green_fill, green_font
                    elif cell.value >= 80:
                        row_fill, row_font = yellow_fill, Font(color="000000")
                    elif cell.value >= 50:
                        row_fill, row_font = orange_fill, Font(color="000000")
                    else:
                        row_fill, row_font = red_fill, red_font

                    for c in range(1, ws.max_column + 1):
                        ws.cell(row, c).fill = row_fill
                        ws.cell(row, c).font = row_font

                    # Keep achievement percentage prominent.
                    cell.font = Font(
                        color=row_font.color.rgb if row_font.color and row_font.color.type == "rgb" else "000000",
                        bold=True
                    )

        if yet_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, yet_col)
                cell.number_format = '#,##0.0'
                if isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.fill, cell.font = red_fill, red_font
                    else:
                        cell.fill, cell.font = green_fill, green_font

        if net_col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, net_col).number_format = '#,##0'

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(vertical="center")

        if ws.title.startswith("Top "):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = green_fill
                    cell.font = green_font
        elif ws.title.startswith("Bottom "):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill, cell.font = red_fill, red_font

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = total_fill
                ws.cell(row, col).font = total_font

    wb.save(path)


def export_excel(data):
    output = Path("Gudivada_Net_Accounts_Latest.xlsx")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data.sort_values(
            ["Office Type", "Achievement %"], ascending=[True, False]
        ).to_excel(writer, index=False, sheet_name="All Offices")

        for office_type_name, sheet in [
            ("BO", "BO Performance"), ("SO", "SO Performance"), ("HO", "HO Performance")
        ]:
            data[data["Office Type"] == office_type_name].sort_values(
                "Achievement %", ascending=False
            ).to_excel(writer, index=False, sheet_name=sheet)

        bo = data[data["Office Type"] == "BO"]
        so = data[data["Office Type"] == "SO"]

        bo.sort_values("Achievement %", ascending=False).head(5).to_excel(
            writer, index=False, sheet_name="Top 5 BOs")
        bo.sort_values("Achievement %").head(5).to_excel(
            writer, index=False, sheet_name="Bottom 5 BOs")
        so.sort_values("Achievement %", ascending=False).head(3).to_excel(
            writer, index=False, sheet_name="Top 3 SOs")
        so.sort_values("Achievement %").head(3).to_excel(
            writer, index=False, sheet_name="Bottom 3 SOs")

        summary = pd.DataFrame({
            "Metric": [
                "Report Date", "Total Offices", "BOs", "SOs", "HOs",
                "Accounts Opened", "Accounts Closed", "Net Accounts"
            ],
            "Value": [
                str(date.today()), len(data),
                int((data["Office Type"] == "BO").sum()),
                int((data["Office Type"] == "SO").sum()),
                int((data["Office Type"] == "HO").sum()),
                data["Accounts Opened"].sum(),
                data["Accounts Closed"].sum(),
                data["Net Accounts"].sum()
            ]
        })
        summary.to_excel(writer, index=False, sheet_name="Summary")

    style_workbook(output)
    return output

def main():
    try:
        messagebox.showinfo(
            "Gudivada Net Accounts Updater v2",
            "Select the latest CEPT ACCOUNTS OPENED report."
        )

        opened = choose_file("Select Accounts OPENED report")
        if not opened:
            return

        closed = choose_file("Select Accounts CLOSED report")
        if not closed:
            return

        data = build_dashboard(opened, closed)
        output = export_excel(data)

        messagebox.showinfo(
            "Completed",
            "Gudivada Net Accounts updated successfully.\n\n"
            f"Offices: {len(data)}\n"
            f"Opened: {data['Accounts Opened'].sum():,.0f}\n"
            f"Closed: {data['Accounts Closed'].sum():,.0f}\n"
            f"Net: {data['Net Accounts'].sum():,.0f}\n\n"
            f"Created:\n{output.resolve()}"
        )

    except Exception as e:
        messagebox.showerror(
            "Update Failed",
            f"{type(e).__name__}: {e}"
        )

if __name__ == "__main__":
    main()
