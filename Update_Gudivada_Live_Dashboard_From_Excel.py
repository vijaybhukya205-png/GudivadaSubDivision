from pathlib import Path
import openpyxl, json, re, html

BASE=Path(__file__).resolve().parent
XLSX=BASE/"Gudivada_Net_Accounts_Latest.xlsx"
INDEX=BASE/"index.html"
wb=openpyxl.load_workbook(XLSX,data_only=True)

def read_rows(sheet):
    ws=wb[sheet]
    h=[c.value for c in ws[1]]
    return [dict(zip(h,r)) for r in ws.iter_rows(min_row=2,values_only=True) if r[0]]

index_text=INDEX.read_text(encoding="utf-8")
m=re.search(r'const D=(\{.*?\});',index_text,re.S)
if not m: raise RuntimeError("const D data block not found in index.html")
old=json.loads(m.group(1))
parent={x["Office Name"]:x.get("Parent SO","") for x in old.get("all",[])}

def enrich(rows):
    out=[]
    for x in rows:
        y=dict(x); y["Parent SO"]=parent.get(y["Office Name"],y.get("Parent SO","")); out.append(y)
    return out

bo=enrich(read_rows("BO Performance"))
so=enrich(read_rows("SO Performance"))
allr=enrich(read_rows("All Offices"))

combined=[]
for s in so:
    children=[b for b in bo if b.get("Parent SO")==s["Office Name"]]
    target=(s.get("Annual Target") or 0)+sum(b.get("Annual Target") or 0 for b in children)
    prop=(s.get("Proportionate Target") or 0)+sum(b.get("Proportionate Target") or 0 for b in children)
    net=(s.get("Net Accounts") or 0)+sum(b.get("Net Accounts") or 0 for b in children)
    combined.append({"SO Name":s["Office Name"],"BOs":len(children),"Combined Target":target,
                     "Proportionate Target":prop,"Net Accounts":net,
                     "Achievement %":(net/prop*100 if prop else 0),"Yet to Achieve":max(prop-net,0)})

data={"bo":bo,"so_own":so,"so_comb":combined,"all":allr}
blob=json.dumps(data,ensure_ascii=False,separators=(",",":"))
index_text=index_text[:m.start()]+"const D="+blob+";"+index_text[m.end():]

report_date=None
try:
    for row in wb["Summary"].iter_rows(min_row=2,values_only=True):
        if row[0]=="Report Date":
            v=row[1]
            if hasattr(v,"strftime"): report_date=v.strftime("%d-%m-%Y")
            break
except Exception: pass
if report_date:
    index_text=re.sub(r'As on Date:\s*\d{2}-\d{2}-\d{4}',f'As on Date: {report_date}',index_text)
INDEX.write_text(index_text,encoding="utf-8")

def make_report(title,rows,cols):
    th="".join(f"<th>{html.escape(a)}</th>" for a,b in cols)
    body=[]
    for i,r in enumerate(sorted(rows,key=lambda x:float(x.get("Achievement %",0)),reverse=True),1):
        p=float(r.get("Achievement %",0)); cls="g" if p>=100 else "y" if p>=80 else "o" if p>=60 else "r"
        cells=[]
        for a,key in cols:
            if key=="__rank__": v=i
            elif key=="__yet_open__": v=max(float(r.get("Proportionate Target",0))-float(r.get("Net Accounts",0)),0)
            else: v=r.get(key,"")
            if key=="Achievement %": v=f"{float(v or 0):.1f}%"
            elif isinstance(v,(int,float)): v=f"{v:,.0f}"
            cells.append(f"<td>{html.escape(str(v))}</td>")
        body.append(f'<tr class="{cls}">{"".join(cells)}</tr>')
    css="""@page{size:A4 landscape;margin:8mm}*{box-sizing:border-box}body{font-family:Arial,sans-serif;margin:0;color:#17202a}
h1{font-size:20px;color:#173f5f;margin:0 0 5px}.meta{font-size:11px;margin-bottom:12px}
table{width:100%;border-collapse:collapse;font-size:8.5px}th{background:#173f5f;color:#fff;padding:5px;border:1px solid #173f5f;text-align:left}
td{padding:4px 5px;border:1px solid #d8dde3}tr.g td{background:#c6efce}tr.y td{background:#ffeb9c}tr.o td{background:#fce4d6}tr.r td{background:#ffc7ce}
.actions{margin-bottom:10px}button{padding:7px 14px;margin-right:6px}@media print{.actions{display:none}body{-webkit-print-color-adjust:exact;print-color-adjust:exact}}"""
    date=report_date or "22-08-2026"
    return f"""<!doctype html><html><head><meta charset="utf-8"><title>{html.escape(title)}</title><style>{css}</style></head>
<body><div class="actions"><button onclick="window.print()">Print / Save PDF</button><button onclick="window.close()">Close</button></div>
<h1>{html.escape(title)}</h1><div class="meta"><b>As on Date: {date}</b> | Gudivada Sub Division</div>
<table><thead><tr>{th}</tr></thead><tbody>{''.join(body)}</tbody></table></body></html>"""

(BASE/"BO_Performance_Report.html").write_text(make_report("Gudivada Sub Division — BO Performance",bo,
[("Rank","__rank__"),("BO Name","Office Name"),("Parent SO","Parent SO"),("Opened","Accounts Opened"),("Closed","Accounts Closed"),("Net","Net Accounts"),("Proportionate Target","Proportionate Target"),("Achievement","Achievement %"),("Yet to Open","__yet_open__")]),encoding="utf-8")
(BASE/"SO_Performance_Report.html").write_text(make_report("Gudivada Sub Division — SO Performance",so,
[("Rank","__rank__"),("SO Name","Office Name"),("Opened","Accounts Opened"),("Closed","Accounts Closed"),("Net","Net Accounts"),("Proportionate Target","Proportionate Target"),("Achievement","Achievement %"),("Yet to Open","__yet_open__")]),encoding="utf-8")
(BASE/"SO_Consolidated_Report.html").write_text(make_report("Gudivada Sub Division — SO-wise Consolidated Performance",combined,
[("Rank","__rank__"),("SO Name","SO Name"),("BOs","BOs"),("Combined Target","Combined Target"),("Proportionate Target","Proportionate Target"),("Net","Net Accounts"),("Achievement","Achievement %"),("Yet to Achieve","Yet to Achieve"),("Yet to Open","__yet_open__")]),encoding="utf-8")

print(f"Updated dashboard and reports: {len(bo)} BOs, {len(so)} SOs, {len(allr)} offices.")
