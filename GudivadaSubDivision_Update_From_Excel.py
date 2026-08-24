from pathlib import Path
import openpyxl, json, re, html

BASE = Path(__file__).resolve().parent
XLSX = BASE / "Gudivada_Net_Accounts_Latest.xlsx"
HTML = BASE / "index.html"

wb = openpyxl.load_workbook(XLSX, data_only=True)

def read_rows(sheet):
    ws = wb[sheet]
    h = [c.value for c in ws[1]]
    return [dict(zip(h, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

text = HTML.read_text(encoding="utf-8")
m = re.search(r'const D=(\{.*?\});', text, re.S)
if not m:
    raise RuntimeError("const D data block not found in index.html")

old = json.loads(m.group(1))
parent = {x["Office Name"]: x.get("Parent SO", "") for x in old.get("all", [])}

def enrich(items):
    out = []
    for x in items:
        y = dict(x)
        y["Parent SO"] = parent.get(y["Office Name"], "")
        out.append(y)
    return out

bo = enrich(read_rows("BO Performance"))
so = enrich(read_rows("SO Performance"))
allr = enrich(read_rows("All Offices"))

combined = []
for s in so:
    children = [b for b in bo if b.get("Parent SO") == s["Office Name"]]
    target = (s.get("Annual Target") or 0) + sum(b.get("Annual Target") or 0 for b in children)
    prop = (s.get("Proportionate Target") or 0) + sum(b.get("Proportionate Target") or 0 for b in children)
    net = (s.get("Net Accounts") or 0) + sum(b.get("Net Accounts") or 0 for b in children)
    combined.append({
        "SO Name": s["Office Name"], "BOs": len(children),
        "Combined Target": target, "Proportionate Target": prop,
        "Net Accounts": net,
        "Achievement %": (net / prop * 100 if prop else 0),
        "Yet to Achieve": max(prop - net, 0)
    })

data = {"bo": bo, "so_own": so, "so_comb": combined, "all": allr}
text = text[:m.start()] + "const D=" + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";" + text[m.end():]

try:
    d = None
    for row in wb["Summary"].iter_rows(min_row=2, values_only=True):
        if row[0] == "Report Date":
            v = row[1]
            if hasattr(v, "strftime"):
                d = v.strftime("%d-%m-%Y")
            else:
                sv = str(v)
                if len(sv) >= 10 and sv[4] == "-" and sv[7] == "-":
                    d = f"{sv[8:10]}-{sv[5:7]}-{sv[:4]}"
            break
    if d:
        text = re.sub(r'As on Date:\s*\d{2}-\d{2}-\d{4}', f'As on Date: {d}', text)
except Exception:
    pass

HTML.write_text(text, encoding="utf-8")
print("LIVE HEALTH CARD UPDATED FROM Gudivada_Net_Accounts_Latest.xlsx")

# ---------------------------------------------------------------------------
# POSB standalone PDF/HTML reports â€” regenerate from the latest workbook
# and add report-specific GRAND TOTAL rows.
# ---------------------------------------------------------------------------
def fmt_int(v):
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return str(v or "")

def fmt_pct(v):
    try:
        return f"{float(v):.1f}%"
    except Exception:
        return "0.0%"

def row_class(ach):
    try:
        a = float(ach)
    except Exception:
        a = 0
    if a >= 100:
        return "g"
    if a >= 80:
        return "y"
    if a >= 60:
        return "o"
    return "r"

def get_report_style(template_name):
    tp = BASE / template_name
    if tp.exists():
        s = tp.read_text(encoding="utf-8", errors="ignore")
        mm = re.search(r"<style>(.*?)</style>", s, re.S | re.I)
        return mm.group(1) if mm else ""
    return ""

def report_html(title, date_text, headers, rows, total_values, template_name):
    style = get_report_style(template_name)
    # Ensure the approved portrait/print settings and total-row styling exist.
    if "@page" not in style:
        style += "@page{size:A4 portrait;margin:10mm}"
    style += """
    .grand-total td{font-weight:800;background:#e8eef5!important;color:#000!important;
      border-top:2px solid #173f5f!important;-webkit-print-color-adjust:exact;
      print-color-adjust:exact}
    @media print{
      @page{size:A4 portrait;margin:10mm}
      th{background:#173f5f!important;color:#fff!important;
        -webkit-print-color-adjust:exact;print-color-adjust:exact}
      .grand-total td{background:#e8eef5!important}
    }
    """
    body = []
    for i, r in enumerate(rows, 1):
        ach = r.get("Achievement %", 0)
        body.append(
            f'<tr class="{row_class(ach)}">'
            f'<td>{i}</td><td>{html.escape(str(r["Office Name"]))}</td>'
            f'<td>{html.escape(str(r.get("Office Type","")))}</td>'
            f'<td>{fmt_int(r.get("Accounts Opened",0))}</td>'
            f'<td>{fmt_int(r.get("Accounts Closed",0))}</td>'
            f'<td>{fmt_int(r.get("Net Accounts",0))}</td>'
            f'<td>{fmt_int(r.get("Proportionate Target",0))}</td>'
            f'<td>{fmt_pct(ach)}</td>'
            f'<td>{fmt_int(r.get("Yet to Achieve",0))}</td></tr>'
        )
    body.append(
        '<tr class="grand-total"><td></td><td>GRAND TOTAL</td><td></td>'
        f'<td>{fmt_int(total_values["opened"])}</td>'
        f'<td>{fmt_int(total_values["closed"])}</td>'
        f'<td>{fmt_int(total_values["net"])}</td>'
        f'<td>{fmt_int(total_values["target"])}</td>'
        f'<td>{fmt_pct(total_values["achievement"])}</td>'
        f'<td>{fmt_int(total_values["yet"])}</td></tr>'
    )
    return f"""<!doctype html><html><head><meta charset="utf-8">
<title>{html.escape(title)}</title><style>{style}</style></head><body>
<div class="actions"><button onclick="window.print()">Print / Save PDF</button>
<button onclick="window.close()">Close</button></div>
<h1>{html.escape(title)}</h1><div class="meta"><b>As on Date: {html.escape(date_text)}</b></div>
<table><thead><tr>{''.join(f'<th>{html.escape(h)}</th>' for h in headers)}</tr></thead>
<tbody>{''.join(body)}</tbody></table></body></html>"""

def rows_from_sheet(sheet_name):
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    result = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        if not vals or not vals[0]:
            continue
        d = dict(zip(headers, vals))
        result.append(d)
    return result

# Report date from Summary.
report_date = ""
for vals in wb["Summary"].iter_rows(min_row=2, values_only=True):
    if vals and vals[0] == "Report Date":
        v = vals[1]
        if hasattr(v, "strftime"):
            report_date = v.strftime("%d-%m-%Y")
        else:
            sv = str(v)
            report_date = (f"{sv[8:10]}-{sv[5:7]}-{sv[:4]}"
                           if len(sv) >= 10 and sv[4] == "-" and sv[7] == "-"
                           else sv)
        break

bo_rows = rows_from_sheet("BO Performance")
so_rows = rows_from_sheet("SO Performance")
ho_rows = rows_from_sheet("HO Performance")

def total_for(rows):
    opened = sum(float(r.get("Accounts Opened") or 0) for r in rows)
    closed = sum(float(r.get("Accounts Closed") or 0) for r in rows)
    net = sum(float(r.get("Net Accounts") or 0) for r in rows)
    target = sum(float(r.get("Proportionate Target") or 0) for r in rows)
    ach = (net / target * 100) if target else 0
    yet = max(target - net, 0)
    return {"opened": opened, "closed": closed, "net": net,
            "target": target, "achievement": ach, "yet": yet}

headers_standard = ["Sl. No.","Office Name","Type","Opened","Closed",
                    "Net Accounts","Proportionate Target","Achievement","Yet to Open"]

# BO report
( BASE / "BO_Performance_Report.html" ).write_text(
    report_html("Gudivada Sub Division - BO Performance Report", report_date,
                headers_standard, bo_rows, total_for(bo_rows),
                "BO_Performance_Report.html"), encoding="utf-8")

# SO report
( BASE / "SO_Performance_Report.html" ).write_text(
    report_html("Gudivada Sub Division - SO Performance Report", report_date,
                headers_standard, so_rows, total_for(so_rows),
                "SO_Performance_Report.html"), encoding="utf-8")

# SO + HO report: only SO + HO rows in this report.
soho_rows = so_rows + ho_rows
soho_rows.sort(key=lambda r: float(r.get("Achievement %") or 0), reverse=True)
( BASE / "SO_HO_Performance_Report.html" ).write_text(
    report_html("Gudivada Sub Division - SO + HO Performance Report", report_date,
                headers_standard, soho_rows, total_for(soho_rows),
                "SO_HO_Performance_Report.html"), encoding="utf-8")

# SO-wise consolidated report:
# Complete Sub-Division consolidation:
#   - Every SO is included, even if it has no mapped BO.
#   - Each SO row includes its own SO target + mapped BO annual targets.
#   - Gudivada HO is included as a separate non-BO office.
#   - Grand Total = 15 SO targets + 72 BO targets + HO target.
parent_map = {x["Office Name"]: x.get("Parent SO","") for x in old.get("all", [])}

# Build BOs under each SO using the existing Dashboard master mapping.
bo_by_so = {}
for b in bo_rows:
    so_name = parent_map.get(b["Office Name"], "")
    if so_name:
        bo_by_so.setdefault(so_name, []).append(b)

consolidated = []
for s in so_rows:
    so_name = s["Office Name"]
    children = bo_by_so.get(so_name, [])

    so_annual = float(s.get("Annual Target") or 0)
    so_prop = float(s.get("Proportionate Target") or 0)
    bo_annual = sum(float(x.get("Annual Target") or 0) for x in children)
    bo_prop = sum(float(x.get("Proportionate Target") or 0) for x in children)

    so_opened = float(s.get("Accounts Opened") or 0)
    so_closed = float(s.get("Accounts Closed") or 0)
    bo_opened = sum(float(x.get("Accounts Opened") or 0) for x in children)
    bo_closed = sum(float(x.get("Accounts Closed") or 0) for x in children)

    combined_annual = so_annual + bo_annual
    combined_prop = so_prop + bo_prop
    combined_opened = so_opened + bo_opened
    combined_closed = so_closed + bo_closed
    combined_net = combined_opened - combined_closed
    ach = combined_net / combined_prop * 100 if combined_prop else 0

    consolidated.append({
        "Office Name": so_name,
        "Office Type": "SO",
        "BOs": len(children),
        "SO Annual Target": so_annual,
        "BO Annual Target": bo_annual,
        "Combined Target": combined_annual,
        "Proportionate Target": combined_prop,
        "Accounts Opened": combined_opened,
        "Accounts Closed": combined_closed,
        "Net Accounts": combined_net,
        "Achievement %": ach,
        "Yet to Achieve": max(combined_prop - combined_net, 0)
    })

# Keep all 15 SOs, ranked by achievement.
consolidated.sort(key=lambda r: float(r.get("Achievement %") or 0), reverse=True)

# HO is a non-BO office and must appear separately.
ho_rows = [r for r in ho_rows if r.get("Office Name")]
ho = ho_rows[0] if ho_rows else None

def consolidated_html():
    style = get_report_style("SO_Consolidated_Report.html")
    style += """
    .grand-total td{font-weight:800;background:#e8eef5!important;color:#000!important;
      border-top:2px solid #173f5f!important;-webkit-print-color-adjust:exact;
      print-color-adjust:exact}
    .office-total td{font-weight:700;background:#f5f7fa!important;
      -webkit-print-color-adjust:exact;print-color-adjust:exact}
    @media print{
      @page{size:A4 portrait;margin:10mm}
      th{background:#173f5f!important;color:#fff!important;
        -webkit-print-color-adjust:exact;print-color-adjust:exact}
      .grand-total td{background:#e8eef5!important}
      .office-total td{background:#f5f7fa!important}
    }
    """

    rows_html = []
    for i, r in enumerate(consolidated, 1):
        rows_html.append(
            f'<tr class="{row_class(r["Achievement %"])}">'
            f'<td>{i}</td>'
            f'<td>{html.escape(str(r["Office Name"]))}</td>'
            f'<td>{r["BOs"]}</td>'
            f'<td>{fmt_int(r["SO Annual Target"])}</td>'
            f'<td>{fmt_int(r["BO Annual Target"])}</td>'
            f'<td>{fmt_int(r["Combined Target"])}</td>'
            f'<td>{fmt_int(r["Proportionate Target"])}</td>'
            f'<td>{fmt_int(r["Net Accounts"])}</td>'
            f'<td>{fmt_pct(r["Achievement %"])}</td>'
            f'<td>{fmt_int(r["Yet to Achieve"])}</td></tr>'
        )

    # HO/non-BO row.
    if ho:
        ho_annual = float(ho.get("Annual Target") or 0)
        ho_prop = float(ho.get("Proportionate Target") or 0)
        ho_net = float(ho.get("Net Accounts") or 0)
        ho_ach = ho_net / ho_prop * 100 if ho_prop else 0
        ho_yet = max(ho_prop - ho_net, 0)
        rows_html.append(
            f'<tr class="office-total">'
            f'<td>-</td><td>{html.escape(str(ho["Office Name"]))}</td>'
            f'<td>â€”</td><td>â€”</td><td>â€”</td>'
            f'<td>{fmt_int(ho_annual)}</td><td>{fmt_int(ho_prop)}</td>'
            f'<td>{fmt_int(ho_net)}</td><td>{fmt_pct(ho_ach)}</td>'
            f'<td>{fmt_int(ho_yet)}</td></tr>'
        )

    # Complete Sub-Division Grand Total:
    # 15 SO x 600 + 72 BO x 300 + HO 2000 = 32600 annual target.
    total_so_annual = sum(float(r.get("SO Annual Target") or 0) for r in consolidated)
    total_bo_annual = sum(float(r.get("BO Annual Target") or 0) for r in consolidated)
    total_so_prop = sum(float(r.get("Proportionate Target") or 0) for r in consolidated) - total_bo_annual * 5 / 12
    total_bo_prop = total_bo_annual * 5 / 12

    # Use actual row targets from the workbook, but force the stated 5-month
    # proportionate basis for the consolidated report.
    elapsed_months = 5
    total_annual = total_so_annual + total_bo_annual + (float(ho.get("Annual Target") or 0) if ho else 0)
    total_prop = total_annual * elapsed_months / 12
    total_net = sum(float(r.get("Net Accounts") or 0) for r in consolidated) + (float(ho.get("Net Accounts") or 0) if ho else 0)
    total_ach = total_net / total_prop * 100 if total_prop else 0
    total_yet = max(total_prop - total_net, 0)
    total_offices = len(consolidated) + len(bo_rows) + (1 if ho else 0)

    rows_html.append(
        f'<tr class="grand-total">'
        f'<td></td><td>GRAND TOTAL</td><td>{total_offices} Offices</td>'
        f'<td>{fmt_int(total_so_annual)}</td>'
        f'<td>{fmt_int(total_bo_annual)}</td>'
        f'<td>{fmt_int(total_annual)}</td>'
        f'<td>{fmt_int(total_prop)}</td>'
        f'<td>{fmt_int(total_net)}</td>'
        f'<td>{fmt_pct(total_ach)}</td>'
        f'<td>{fmt_int(total_yet)}</td></tr>'
    )

    return f"""<!doctype html><html><head><meta charset="utf-8">
<title>Gudivada Sub Division - SO-wise Consolidated Report</title><style>{style}</style></head><body>
<div class="actions"><button onclick="window.print()">Print / Save PDF</button>
<button onclick="window.close()">Close</button></div>
<h1>Gudivada Sub Division - SO-wise Consolidated Report</h1>
<div class="meta"><b>As on Date: {html.escape(report_date)}</b></div>
<table><thead><tr>
<th>Sl. No.</th><th>SO / Office Name</th><th>BOs</th>
<th>SO Annual Target</th><th>BO Annual Target</th><th>Combined Target</th>
<th>Proportionate Target</th><th>Net Accounts</th><th>Achievement</th><th>Yet to Open</th>
</tr></thead>
<tbody>{''.join(rows_html)}</tbody></table></body></html>"""

(BASE / "SO_Consolidated_Report.html").write_text(consolidated_html(), encoding="utf-8")
print("POSB reports regenerated with report-specific GRAND TOTAL rows.")


